#!/usr/bin/env python3
"""
Create sample features.xlsx for AI Agent testing
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Features"
    
    # Add headers
    headers = ["ID", "Feature", "Description", "Priority"]
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data
    features = [
        (1, "User Authentication", "Implement user login with JWT tokens", "High"),
        (2, "QR History", "Store generated QR codes in database", "High"),
        (3, "Analytics Dashboard", "Display usage statistics and trends", "Medium"),
        (4, "API Rate Limiting", "Prevent abuse with rate limiting", "Medium"),
        (5, "Export Features", "Allow exporting QR as SVG/PDF formats", "Low"),
        (6, "Batch Generation", "Generate multiple QR codes at once", "Low"),
        (7, "Custom Branding", "Allow logo embedding in QR codes", "Low"),
    ]
    
    for feature in features:
        ws.append(feature)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    
    # Center align priority column
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=len(features)+1):
        row[3].alignment = Alignment(horizontal="center")
    
    # Save file
    wb.save("features.xlsx")
    print("✓ Sample features.xlsx created successfully!")
    
except ImportError:
    print("openpyxl not installed. Creating sample.txt instead...")
    
    # Create text version
    with open("features.txt", "w") as f:
        f.write("Sample Requirements\n")
        f.write("==================\n\n")
        f.write("1. User Authentication - HIGH PRIORITY\n")
        f.write("   Implement user login with JWT tokens and session management.\n\n")
        f.write("2. QR History - HIGH PRIORITY\n")
        f.write("   Store generated QR codes in database for tracking.\n\n")
        f.write("3. Analytics Dashboard - MEDIUM PRIORITY\n")
        f.write("   Display usage statistics and trends over time.\n\n")
        f.write("4. API Rate Limiting - MEDIUM PRIORITY\n")
        f.write("   Prevent abuse with rate limiting per IP.\n\n")
        f.write("5. Export Features - LOW PRIORITY\n")
        f.write("   Allow exporting QR as SVG/PDF formats.\n\n")
    
    print("✓ Sample features.txt created successfully!")

except Exception as e:
    print(f"Error creating sample: {e}")
